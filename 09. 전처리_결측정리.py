"""
B 파일: 삭제 + 0 대체 + 구조 정리 + 로그 출력
=====================================================
역할
1) price / assets / equity 모두 NaN인 행 삭제
2) 지정 컬럼 0 대체
3) div_yield 제거, ticker 포맷 정리, 헤더 중복 행 제거
4) 삭제 로그 엑셀 저장

주의
- interest_coverage 조건부 NaN 처리 같은 이상치/논리 처리 규칙은 하지 않음
- 그 작업은 C 파일에서 수행
"""

from __future__ import annotations

import os
from datetime import datetime
from typing import Dict, List, Tuple

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# ============================================================
# 설정
# ============================================================
DATASETS = [
    {
        "label": "KOSPI",
        "input": "master_kospi.csv",
        "output": "master_kospi_preclean.csv",
    },
    {
        "label": "KOSDAQ",
        "input": "master_kosdaq.csv",
        "output": "master_kosdaq_preclean.csv",
    },
]

LOG_FILE = "preclean_drop_log.xlsx"
CORE_DROP_COLS = ["price", "assets", "equity"]
ZERO_FILL_COLS = [
    "capital_increase",
    "treasury",
    "short_liab",
    "dividend",
    "div_ratio",
    "interest",
]
DROP_COLS_IF_EXISTS = ["div_yield"]

# 로그에 우선 실을 컬럼
LOG_COLS = [
    "quarter",
    "period",
    "ticker",
    "corp_name",
    "sector",
    "price",
    "assets",
    "equity",
    "shares",
]

HEADER_TOKENS = {"quarter", "ticker", "corp_name", "sector", "period", "종목코드", "기업명", "분기"}


# ============================================================
# 유틸
# ============================================================
def load_table(path: str) -> pd.DataFrame:
    ext = os.path.splitext(path)[1].lower()
    if ext in {".xlsx", ".xls"}:
        return pd.read_excel(path, dtype={"ticker": str})
    if ext == ".csv":
        try:
            return pd.read_csv(path, dtype={"ticker": str}, encoding="utf-8-sig")
        except UnicodeDecodeError:
            return pd.read_csv(path, dtype={"ticker": str}, encoding="cp949")
    raise ValueError(f"지원하지 않는 파일 형식: {path}")


def save_table(df: pd.DataFrame, path: str) -> None:
    ext = os.path.splitext(path)[1].lower()
    if ext in {".xlsx", ".xls"}:
        df.to_excel(path, index=False)
        return
    if ext == ".csv":
        df.to_csv(path, index=False, encoding="utf-8-sig")
        return
    raise ValueError(f"지원하지 않는 저장 형식: {path}")


def normalize_ticker(df: pd.DataFrame) -> pd.DataFrame:
    if "ticker" in df.columns:
        df["ticker"] = df["ticker"].astype(str).str.strip().str.zfill(6)
    return df


def drop_header_like_rows(df: pd.DataFrame) -> Tuple[pd.DataFrame, int]:
    if df.empty:
        return df, 0

    mask = pd.Series(False, index=df.index)
    for col in [c for c in ["ticker", "quarter", "period", "corp_name"] if c in df.columns]:
        as_str = df[col].astype(str).str.strip()
        mask = mask | as_str.eq(col) | as_str.isin(HEADER_TOKENS)

    removed = int(mask.sum())
    if removed:
        df = df.loc[~mask].reset_index(drop=True)
    return df, removed


# ============================================================
# 전처리 본체
# ============================================================
def preprocess_dataset(cfg: Dict[str, str]) -> Dict[str, object]:
    label = cfg["label"]
    input_path = cfg["input"]
    output_path = cfg["output"]

    if not os.path.exists(input_path):
        raise FileNotFoundError(f"[{label}] 입력 파일 없음: {input_path}")

    df = load_table(input_path)
    original_rows = len(df)
    print(f"\n{'=' * 68}")
    print(f"[{label}] 로드 완료: {input_path}")
    print(f"  원본 행 수: {original_rows}")

    df, header_removed = drop_header_like_rows(df)
    if header_removed:
        print(f"  헤더 중복 행 제거: {header_removed}행")

    df = normalize_ticker(df)

    missing_core = [c for c in CORE_DROP_COLS if c not in df.columns]
    if missing_core:
        raise KeyError(f"[{label}] 삭제 조건 컬럼 없음: {missing_core}")

    drop_mask = df[CORE_DROP_COLS].isna().all(axis=1)
    drop_df = df.loc[drop_mask].copy()
    clean_df = df.loc[~drop_mask].reset_index(drop=True)

    print(f"  상장 전/분석불가 행 삭제: {int(drop_mask.sum())}행")

    zero_fill_report: List[Tuple[str, int]] = []
    for col in ZERO_FILL_COLS:
        if col in clean_df.columns:
            cnt = int(clean_df[col].isna().sum())
            clean_df[col] = clean_df[col].fillna(0)
            zero_fill_report.append((col, cnt))

    removed_cols = []
    for col in DROP_COLS_IF_EXISTS:
        if col in clean_df.columns:
            clean_df = clean_df.drop(columns=[col])
            removed_cols.append(col)

    save_table(clean_df, output_path)

    print(f"  저장 완료: {output_path}")
    if zero_fill_report:
        print("  0 대체:")
        for col, cnt in zero_fill_report:
            print(f"    - {col}: {cnt}건")
    if removed_cols:
        print(f"  구조 정리: 컬럼 제거 {removed_cols}")

    return {
        "label": label,
        "input": input_path,
        "output": output_path,
        "original_rows": original_rows,
        "header_removed": header_removed,
        "drop_rows": int(drop_mask.sum()),
        "clean_rows": len(clean_df),
        "drop_df": drop_df,
        "clean_df": clean_df,
        "zero_fill_report": zero_fill_report,
        "removed_cols": removed_cols,
    }


# ============================================================
# 로그 저장
# ============================================================
def style_sheet_header(ws, row: int, values: List[str], header_fill, header_font, border, center):
    for ci, value in enumerate(values, 1):
        cell = ws.cell(row=row, column=ci, value=value)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = center


def autosize_worksheet(ws, max_width: int = 35):
    for col in ws.columns:
        max_len = max((len(str(c.value)) if c.value is not None else 0 for c in col), default=8)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, max_width)


def build_log(results: List[Dict[str, object]], log_path: str) -> None:
    wb = Workbook()
    ws_summary = wb.active
    ws_summary.title = "요약"

    header_fill = PatternFill("solid", start_color="2F5496")
    sub_fill = PatternFill("solid", start_color="D9E1F2")
    warn_fill = PatternFill("solid", start_color="FFE699")
    header_font = Font(name="Arial", bold=True, color="FFFFFF", size=10)
    body_font = Font(name="Arial", size=10)
    center = Alignment(horizontal="center", vertical="center")
    left = Alignment(horizontal="left", vertical="center")
    thin = Side(style="thin", color="BFBFBF")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    summary_rows = [[
        "실행 일시",
        datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
    ]]
    for result in results:
        summary_rows.extend([
            [f"{result['label']} 입력", result["input"]],
            [f"{result['label']} 출력", result["output"]],
            [f"{result['label']} 원본 행 수", result["original_rows"]],
            [f"{result['label']} 헤더 중복 제거", result["header_removed"]],
            [f"{result['label']} 삭제 행 수", result["drop_rows"]],
            [f"{result['label']} 잔여 행 수", result["clean_rows"]],
            [f"{result['label']} 삭제 종목 수", int(result["drop_df"]["ticker"].nunique()) if "ticker" in result["drop_df"].columns else 0],
        ])

    ws_summary.column_dimensions["A"].width = 28
    ws_summary.column_dimensions["B"].width = 44
    for i, (k, v) in enumerate(summary_rows, 1):
        ka = ws_summary.cell(row=i, column=1, value=k)
        va = ws_summary.cell(row=i, column=2, value=v)
        for cell in (ka, va):
            cell.font = body_font
            cell.border = border
            cell.alignment = left
        ka.fill = sub_fill

    for result in results:
        label = result["label"]
        drop_df: pd.DataFrame = result["drop_df"]

        # 삭제행상세 시트
        ws_detail = wb.create_sheet(f"{label}_삭제행상세"[:31])
        detail_cols = [c for c in LOG_COLS if c in drop_df.columns]
        detail_df = drop_df[detail_cols].copy() if detail_cols else drop_df.copy()
        sort_cols = [c for c in ["ticker", "quarter", "period"] if c in detail_df.columns]
        if sort_cols:
            detail_df = detail_df.sort_values(sort_cols).reset_index(drop=True)

        style_sheet_header(ws_detail, 1, list(detail_df.columns), header_fill, header_font, border, center)
        for ri, (_, row) in enumerate(detail_df.iterrows(), start=2):
            for ci, val in enumerate(row, start=1):
                cell = ws_detail.cell(row=ri, column=ci, value=val)
                cell.font = body_font
                cell.border = border
                cell.alignment = center if ci > 1 else left
                cell.fill = warn_fill
        autosize_worksheet(ws_detail, max_width=30)

        # 종목별요약 시트
        ws_ticker = wb.create_sheet(f"{label}_종목별요약"[:31])
        if not drop_df.empty and all(c in drop_df.columns for c in ["ticker", "corp_name"]):
            period_col = "quarter" if "quarter" in drop_df.columns else ("period" if "period" in drop_df.columns else None)
            if period_col:
                ticker_summary = (
                    drop_df.groupby(["ticker", "corp_name"])[period_col]
                    .apply(lambda x: ", ".join(sorted(x.astype(str))))
                    .reset_index()
                    .rename(columns={period_col: "삭제된_분기"})
                )
            else:
                ticker_summary = drop_df[["ticker", "corp_name"]].drop_duplicates().copy()
                ticker_summary["삭제된_분기"] = ""

            ticker_summary["삭제_분기수"] = ticker_summary["삭제된_분기"].apply(
                lambda x: 0 if not str(x).strip() else str(x).count(",") + 1
            )
        else:
            ticker_summary = pd.DataFrame(columns=["ticker", "corp_name", "삭제된_분기", "삭제_분기수"])

        style_sheet_header(ws_ticker, 1, list(ticker_summary.columns), header_fill, header_font, border, center)
        for ri, (_, row) in enumerate(ticker_summary.iterrows(), start=2):
            for ci, val in enumerate(row, start=1):
                cell = ws_ticker.cell(row=ri, column=ci, value=val)
                cell.font = body_font
                cell.border = border
                cell.alignment = center if ci != 2 else left
        autosize_worksheet(ws_ticker, max_width=35)

    wb.save(log_path)
    print(f"\n로그 저장 완료: {log_path}")


# ============================================================
# 실행
# ============================================================
def main():
    results = [preprocess_dataset(cfg) for cfg in DATASETS]
    build_log(results, LOG_FILE)

    print(f"\n{'=' * 68}")
    print("최종 요약")
    print(f"{'=' * 68}")
    for result in results:
        print(
            f"[{result['label']}] {result['input']} -> {result['output']} | "
            f"삭제 {result['drop_rows']}행 | 잔여 {result['clean_rows']}행"
        )


if __name__ == "__main__":
    main()
